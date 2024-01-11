import os, sys, re
from concurrent.futures import ProcessPoolExecutor, as_completed
from multiprocessing import freeze_support
from threading import Thread
from copy import deepcopy
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
import base64
import ujson as json
import datetime, time
import csv
from argparse import ArgumentParser
from azure.iot.hub import IoTHubRegistryManager

# import azure.storage.filedatalake

datablock_config_id = ["Agent_Controls", "Core_File", "Device_Info", "DHCP_Lease_Table", "Domain_Reputation",
                       "DR_Report", "Ethernet_Interface", "IPTV_Info", "Log", "Neighbors", "PBX", "Priorization",
                       "Radar_Detection", "Roaming_Event", "Stations_OnDemand", "Stations_Periodic", "Steering_Event",
                       "VoIP_Info", "WAN_Connection", "Watchdog_Config", "Watchdog_Periodic_Info_DNS",
                       "Watchdog_Restart_Service", "Watchdog_Robustness", "WiFi_Interface"]

datablock_id = ["WiFi_Interface", "Ethernet_Interface", "Stations_Periodic", "Stations_OnDemand",
                "Neighbors", "Device_Info", "WAN_Connection", "Multicast", "VoIP_Info", "IPTV_Info",
                "MAC_Access_Control",
                "IPTABLES", "TX_POWER_TABLE", "Radar_Detection", "DHCP_Lease_Table", "Watchdog_Config",
                "Watchdog_ARP_Table", "Watchdog_IP_Route",
                "Watchdog_Periodic_Info_General", "Watchdog_Periodic_Info_Connections",
                "Watchdog_Periodic_Info_Processes",
                "Watchdog_Periodic_Info_DNS", "Watchdog_Restart_Process", "Watchdog_Alarm_PPP", "Watchdog_Alarm_TR069",
                "Watchdog_Boot", "Watchdog_GPON", "Watchdog_CLI",
                "Watchdog_Max_Time_Exceeded", "Watchdog_VoIP_Registration_Alarm", "Watchdog_VoIP_DHCP_Alarm",
                "Agent_Controls", "Watchdog_Restart_Service", "Watchdog_Robustness", "Domain_Reputation", "DR_Report",
                "PBX",
                "Log", "Roaming_Event", "Steering_Event", "Priorization", "Core_File", "HPNA", "HPNA_Diagnostics",
                "Measurements_Quality", "Measurements_SpeedTest", "Measurement_Config"]

service = [{"methodName": "Reboot", "payload": {}},
           {"methodName": "FactoryReset", "payload": {"type": ["Hard", "Soft"]}},
           {"methodName": "WiFiScan", "payload": {"type": ["Complete", "Soft"], "interface": ["2G", "5G", "All"]}},
           {"methodName": "MACAccessControl",
            "payload": {"type": ["blacklist", "whitelist"], "action": ["add", "remove"], "MAC": "input"}},
           {"methodName": "ResetWiFiDriver", "payload": {"interface": ["2G", "5G"]}},
           {"methodName": "ResetCounters", "payload": {"interface": ["2G", "5G", "LAN", "6", "3", "2"]}},
           {"methodName": "CCAMeasurement", "payload": {"interface": ["2G", "5G"],
                                                        "channel": [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 36, 40,
                                                                    44, 48, 52, 56, 60, 64, 100, 104, 108, 112, 116,
                                                                    120, 124, 128, 132, 136, 140, 144, 149, 153, 157,
                                                                    161, 165]}},
           {"methodName": "ResetService",
            "payload": {"service": ["Internet", "TR069", "IPTV", "VOIP", "DR", "PBX", "OPAH"]}},
           {"methodName": "ResetProcess", "payload": {"service": "input"}},
           {"methodName": "Deauthenticate", "payload": {"MAC": "input"}},
           {"methodName": "GetDatablock", "payload": {"datablock": datablock_id}},
           {"methodName": "execCLI", "payload": {"cli": "input", "shellType": ["main", "regular", "quantenna"]}},
           {"methodName": "LogFile",
            "payload": {"module": ["DomainReputation", "Maui", "Log_Supervision"], "start_interval": "input"}},
           {"methodName": "ScriptMngt",
            "payload": {"action": ["create", "delete"], "type": ["persistent", "session"]
                , "filename": "input", "path": "input", "content": "input"}},
           {"methodName": "ExecScript",
            "payload": {"action": ["execute", "schedule", "remove_schedule"], "type": ["persistent", "session"],
                        "shellType": ["main", "regular", "quantenna"],
                        "filename": "input", "path": "input", "scheduleType": "input"}},
           {"methodName": "HPNADiagnostic", "payload": {"NumPacketsInBurst": "input", "BurstInterval": "input",
                                                        "TestPacketPayloadLength": "input", "PayloadEncoding": "input",
                                                        "PayloadDataGen": "input",
                                                        "PayloadType": ["Pattern", "IncrementByte"],
                                                        "PriorityLevel": "input"}}]

source = ["PQA", "TEF"]

PQA_CONNECTION_STRING = ""
PQA_DATA_URL = ""
PQA_DATA_CREDENTIAL = ""
TEF_CONNECTION_STRING = ""
TEF_DATA_URL = ""
TEF_DATA_CREDENTIAL = ""

valid_column = ['Data_Type', 'DeviceId', 'Internet_WAN_IP', 'Timestamp_Report', 'Datablock_Timestamp', 'Datablock_id']

# deviceID_list = {}
report_data = {}

class main_app(tk.Tk):
    def __init__(self, top=None):
        # V1.2 Add device ID in datalak app
        # V1.21
        # V1.22
        # V1.23 Modify timestamp query rule
        # V1.24 (2021/08/23) Remove get twin after update twin
        # V1.25 (2021/10/13) Add type in ExecScript
        # V1.26 (2021/12/20) Add HPNA function, Add CLI mode
        # V1.261 (2021/12/24) Add ignore null Datablock
        # V1.27 (2022/1/4) Add multiple thread for download azure file and generate excel file
        # V1.271 (2022/1/14) Remove multiple thread due to the performance degrade
        # V1.272 (2022/4/1) based on ed14 spec to update direct method content
        # V1.273 (2022/5/3) convert the incorrect timestamp to zero
        # V1.274 (2023/7/18) add speedtest datablock "Measurements_Quality", "Measurements_SpeedTest", "Measurement_Config"
        super().__init__()
        self.geometry("900x800+400+100")
        self.title("Azure Utility v1.274 (2023/7/18)")

        self.notebook = ttk.Notebook(self)
        self.datalake_frame = datalake_app(self.notebook)
        self.direct_method_frame = direct_method_app(self.notebook)
        self.twin_update_frame = twin_update_app(self.notebook)

        self.notebook.add(self.datalake_frame, text='Download Telemetry/Twin')
        self.notebook.add(self.direct_method_frame, text='Direct Method')
        self.notebook.add(self.twin_update_frame, text='Twin Update')
        self.notebook.select(self.datalake_frame)
        self.notebook.pack(expand=1, fill="both")

    def run(self):
        print("")


class direct_method_app(ttk.Frame):
    def __init__(self, top=None):
        super().__init__()
        loc_x = 0.01
        loc_y = 0.01
        loc_x_gap = 0.1
        loc_y_gap = 0.1
        obj_height = 15
        obj_width = 15

        self.first_payload = True
        # self.geometry("700x700+450+200")
        # self.title("Azure Direct Method Utility")

        self.InputFrame = tk.Frame(self)
        self.InputFrame.place(relx=0.01, rely=0.02, relheight=0.4, relwidth=0.98)
        self.InputFrame.configure(relief='groove')
        self.InputFrame.configure(borderwidth="2")

        self.ControlFrame = tk.Frame(self)
        self.ControlFrame.place(relx=0.01, rely=0.41, relheight=0.3, relwidth=0.98)
        self.ControlFrame.configure(relief='groove')
        self.ControlFrame.configure(borderwidth="2")

        self.OutputFrame = tk.Frame(self)
        self.OutputFrame.place(relx=0.01, rely=0.63, relheight=0.35, relwidth=0.98)
        self.OutputFrame.configure(relief='groove')
        self.OutputFrame.configure(borderwidth="2")

        # InputFrame
        self.ServerLabel = tk.Label(self.InputFrame, text='Server:', anchor='w', justify='left')
        self.ServerLabel.place(relx=0.01, rely=0.05, height=obj_height, width=100)

        self.ServerList = ttk.Combobox(self.InputFrame, width=15, values=source)
        self.ServerList.current(0)
        self.ServerList.place(relx=0.10, rely=0.05, relheight=0.1, relwidth=0.35)

        self.DeviceIDLabel = tk.Label(self.InputFrame, text='DeviceID:', anchor='w', justify='left')
        self.DeviceIDLabel.place(relx=0.01, rely=0.2, height=15, width=100)

        self.DeviceIDText = tk.Entry(self.InputFrame, width=100)
        self.DeviceIDText.place(relx=0.10, rely=0.2, relheight=0.1, relwidth=0.35)
        self.DeviceIDText.insert(tk.END, "cc5d4e53225005b5c907bf610e1a7d8c")

        self.MethodLabel = tk.Label(self.InputFrame, text='Method:', anchor='w', justify='left')
        self.MethodLabel.place(relx=0.01, rely=0.35, height=15, width=100)

        self.method_list = []
        for method in service:
            self.method_list.append(method["methodName"])

        self.MethodList = ttk.Combobox(self.InputFrame, width=100, values=self.method_list)
        self.MethodList.current(0)
        self.MethodList.place(relx=0.10, rely=0.35, relheight=0.1, relwidth=0.35)
        self.MethodList.bind("<<ComboboxSelected>>", self.generate_menu)

        self.PayloadLabel = tk.Label(self.InputFrame, text='Payload:', anchor='w', justify='left')
        self.PayloadLabel.place(relx=0.01, rely=0.50, height=15, width=100)

        self.generate_menu("")

        self.GenerateButton = tk.Button(self.InputFrame, text='Generate', anchor='w', justify='left',
                                        command=self.generate_payload)
        self.GenerateButton.place(relx=0.10, rely=0.62, height=30, width=100)

        self.LoopLabel = tk.Label(self.ControlFrame, text='Loop:', anchor='w', justify='left')
        self.LoopLabel.place(relx=0.01, rely=0.12, height=obj_height, width=100)
        self.LoopText = tk.Entry(self.ControlFrame, width=100)
        self.LoopText.place(relx=0.10, rely=0.12, relheight=0.15, relwidth=0.25)
        self.LoopText.insert(tk.END, "1")

        self.PauseLabel = tk.Label(self.ControlFrame, text='Pause:', anchor='w', justify='left')
        self.PauseLabel.place(relx=0.01, rely=0.3, height=obj_height, width=100)
        self.PauseText = tk.Entry(self.ControlFrame, width=100)
        self.PauseText.place(relx=0.10, rely=0.3, relheight=0.15, relwidth=0.25)
        self.PauseText.insert(tk.END, "1")

        self.SendButton = tk.Button(self.ControlFrame, text='Send', anchor='w', justify='left',
                                    command=self.send_payload)
        self.SendButton.place(relx=0.10, rely=0.5, height=30, width=100)

        self.CountLabel = tk.Label(self.ControlFrame, text='Count:', anchor='w', justify='left')
        self.CountLabel.place(relx=0.4, rely=0.12, height=obj_height, width=100)
        self.CountText = tk.Entry(self.ControlFrame, width=100)
        self.CountText.place(relx=0.55, rely=0.12, relheight=0.15, relwidth=0.25)
        self.CountText.insert(tk.END, "1")

        # OutputFrame
        self.SendText = tk.Text(self.OutputFrame, font=("Helvetica", 8))
        self.SendText.place(relx=0.01, rely=0.01, height=300, width=400)
        self.SendText.insert(tk.END, json.dumps({}, indent=4, sort_keys=False))

        self.StatusText = tk.Entry(self.OutputFrame, font=("Helvetica", 8))
        self.StatusText.place(relx=0.49, rely=0.01, height=25, width=400)

        # Vertical (y) Scroll Bar
        self.ScrollBar = tk.Scrollbar(self.OutputFrame)
        self.ScrollBar.pack(side=tk.RIGHT, fill=tk.Y)
        # Configure the scrollbars
        self.ResponseText = tk.Text(self.OutputFrame, yscrollcommand=self.ScrollBar.set, font=("Helvetica", 8))
        self.ResponseText.place(relx=0.49, rely=0.15, height=300, width=400)
        self.ScrollBar.config(command=self.ResponseText.yview)

    def send_payload(self):
        server = self.ServerList.get()
        device_id = self.DeviceIDText.get()
        method_name = self.MethodList.get()
        method_payload = self.SendText.get("1.0", tk.END)
        method_payload_json = json.loads(method_payload)
        method_payload = json.dumps(method_payload_json)
        # method_payload = json.dumps(method_payload_json['payload'])
        loop_count = int(self.LoopText.get())
        pause_time = int(self.PauseText.get())
        self.ResponseText.delete("1.0", tk.END)
        self.SendButton.config(state=tk.DISABLED)

        for count in range(0, loop_count):
            self.CountText.delete(0, tk.END)
            self.CountText.insert(tk.END, str(count + 1))
            self.ResponseText.insert(tk.END, "Test Count:" + str(count + 1) + "\n")
            self.thread_app = send_direct_method(self, server, device_id, method_name, method_payload)
            self.after(pause_time * 1000, "")
            self.CountText.update()
            # app.update()
            app.update_idletasks()
            # self.thread_app.join()
        self.SendButton.config(state=tk.NORMAL)

        return True

    def generate_payload(self):
        payload_text = ""
        service_dict = {}
        selected_name = self.MethodList.get()
        for method in service:
            if method['methodName'] == selected_name:
                service_dict = deepcopy(method)
                for para in service_dict['payload']:
                    for obj in self.payload_obj_list:
                        if para in obj:
                            try:
                                if para == "channel" or para == "start_interval":
                                    service_dict['payload'][para] = int(obj[para].get())
                                else:
                                    service_dict['payload'][para] = obj[para].get().replace("\\n", "\n")
                            except:
                                print("Error")
        payload_text = json.dumps(service_dict['payload'], indent=1, sort_keys=False)
        # payload_text = json.dumps(service_dict, indent=4, sort_keys=False)

        self.SendText.delete("1.0", tk.END)
        self.SendText.insert(tk.END, payload_text)
        return payload_text

    def generate_menu(self, event):
        self.label_arry = []
        self.payload_data_arry = []
        selected_name = self.MethodList.get()
        # print(self.MethodList.get())
        # if self.MethodList.get()=="GetDatablock":
        #    print("Error")
        loc_x = 0.46
        loc_y = 0.05
        if self.first_payload == False:
            for item in self.payload_obj_list:
                for para in item:
                    item[para].destroy()
            for item in self.payload_label_list:
                for para in item:
                    item[para].destroy()
            self.payload_obj_list = []
            self.payload_label_list = []
        else:
            self.payload_obj_list = []
            self.payload_label_list = []

        for method in service:
            if method['methodName'] == selected_name:
                if len(method['payload']) == 0:
                    obj_label = tk.Label(self.InputFrame, text="Content:", anchor='w', justify='left')
                    obj_label.place(relx=loc_x, rely=loc_y, height=15, width=100)
                    obj_payload = tk.Entry(self.InputFrame, width=100)
                    obj_payload.insert(tk.END, "{}")
                    obj_payload.place(relx=loc_x + 0.13, rely=loc_y - 0.02, relheight=0.1, relwidth=0.4)
                    loc_y += 0.1
                    self.payload_label_list.append({'payload': obj_label})
                    self.payload_obj_list.append({'payload': obj_payload})
                else:
                    for para in method['payload']:
                        if type(method['payload'][para]) == list:

                            obj_label = tk.Label(self.InputFrame, text=para + ":", anchor='w', justify='left')
                            obj_label.place(relx=loc_x, rely=loc_y, height=15, width=185)
                            obj_payload = ttk.Combobox(self.InputFrame, width=100, values=method['payload'][para])
                            obj_payload.current(0)
                            obj_payload.place(relx=loc_x + 0.22, rely=loc_y - 0.02, relheight=0.1, relwidth=0.3)
                            loc_y += 0.12
                        else:
                            obj_label = tk.Label(self.InputFrame, text=para + ":", anchor='w', justify='left')
                            obj_label.place(relx=loc_x, rely=loc_y, height=15, width=185)
                            obj_payload = tk.Entry(self.InputFrame, width=100)
                            obj_payload.place(relx=loc_x + 0.22, rely=loc_y - 0.02, relheight=0.1, relwidth=0.3)
                            if para == "MAC":
                                obj_payload.insert(tk.END, "XX:XX:XX:XX:XX:XX")
                            loc_y += 0.12
                        self.payload_label_list.append({para: obj_label})
                        self.payload_obj_list.append({para: obj_payload})
        self.first_payload = False


class send_direct_method(Thread):
    def __init__(self, app, server, device_id, method_name, method_payload):
        super().__init__()

        self.app = app
        self.server = server
        self.device_id = device_id
        self.method_name = method_name
        self.method_payload = method_payload
        # self.app.ResponseText.delete("1.0", tk.END)
        self.start_time = time.time()
        self.start()

    def run(self):
        from azure.iot.hub.models import CloudToDeviceMethod
        try:
            # Create IoTHubRegistryManager
            connection = ""
            if self.server == "PQA":
                connection = PQA_CONNECTION_STRING
            else:
                connection = TEF_CONNECTION_STRING
            registry_manager = IoTHubRegistryManager(connection)

            # Call the direct method.
            # device = {self.device_id}

            # params = {'connectTimeoutInSeconds':30, 'responseTimeoutInSeconds':30}
            # print("--- %s seconds ---" % (time.time() - self.start_time))
            self.method_payload = json.loads(self.method_payload)
            deviceMethod = CloudToDeviceMethod(method_name=self.method_name, payload=self.method_payload,
                                               connect_timeout_in_seconds=30, response_timeout_in_seconds=30)
            response = registry_manager.invoke_device_method(self.device_id, deviceMethod)
            # print("--- %s seconds ---" % (time.time() - self.start_time))
            # print(response.payload)

            # cli_output
            # log_content
            # response

            if "cli_output" in response.payload:
                response.payload['cli_output'] = base64.b64decode(response.payload['cli_output'], validate=True).decode(
                    "utf-8")
            elif "log_content" in response.payload:
                response.payload['log_content'] = base64.b64decode(response.payload['log_content'],
                                                                   validate=True).decode("utf-8")
            elif "response" in response.payload:
                if response.payload['response'] != "Script Scheduled":
                    response.payload['response'] = base64.b64decode(response.payload['response'], validate=True).decode(
                        "utf-8")
                else:
                    response.payload['response'] = response.payload['response']

            text = json.dumps(response.payload, indent=4, sort_keys=False)
            # print("--- %s seconds ---" % (time.time() - self.start_time))
            # self.app.ResponseText.delete("1.0", tk.END)
            if len(sys.argv) == 1:
                self.app.ResponseText.insert(tk.END, text + "\n")
                self.app.ResponseText.update()

                self.app.StatusText.delete(0, tk.END)
                self.app.StatusText.insert(tk.END, "Status Code: " + str(response.status))
                self.app.StatusText.update()

                self.app.update()
                self.app.update_idletasks()
            else:
                print("Status Code:", str(response.status))
                print("Output Text", text)


        #     while True:
        #         print("")
        #         print("IoTHubClient waiting for commands, press Ctrl-C to exit")
        #
        #         status_counter = 0
        #         while status_counter <= WAIT_COUNT:
        #             twin_info = registry_manager.get_twin(DEVICE_ID)
        #
        #             if twin_info.properties.reported.get("rebootTime") != None:
        #                 print("Last reboot time: " + twin_info.properties.reported.get("rebootTime"))
        #             else:
        #                 print("Waiting for device to report last reboot time...")
        #
        #             time.sleep(5)
        #             status_counter += 1
        #
        except Exception as ex:
            if len(sys.argv) == 1:
                text = "Unexpected error {0}".format(ex)
                # self.app.ResponseText.delete("1.0", tk.END)
                self.app.ResponseText.insert(tk.END, text + "\n")
                self.app.ResponseText.update()
                self.app.update()
                self.app.update_idletasks()
            else:
                print("Unexpected error {0}".format(ex))
        return


class datalake_app(ttk.Frame):
    def __init__(self, top=None):
        super().__init__()
        self.ServerLabel = tk.Label(self, text='Server:', anchor='w', justify='left')
        self.ServerLabel.place(relx=0.01, rely=0.05, height=15, width=100)

        self.ServerList = ttk.Combobox(self, width=5, values=source)
        self.ServerList.current(0)
        self.ServerList.place(relx=0.13, rely=0.05, relheight=0.03, relwidth=0.35)

        self.DateLabel = tk.Label(self, text='Date:', anchor='w', justify='left')
        self.DateLabel.place(relx=0.01, rely=0.1, height=15, width=100)

        date_select = ['today', 'all', 'Input_Date']
        self.DateList = ttk.Combobox(self, width=5, values=date_select)
        self.DateList.current(0)
        self.DateList.place(relx=0.13, rely=0.1, relheight=0.03, relwidth=0.35)

        self.DeviceIDLabel = tk.Label(self, text='DeviceID:', anchor='w', justify='left')
        self.DeviceIDLabel.place(relx=0.01, rely=0.15, height=15, width=100)

        deviceID_select = ['all', 'Input_DeviceID']
        self.DeviceIDList = ttk.Combobox(self, width=5, values=deviceID_select)
        self.DeviceIDList.current(0)
        self.DeviceIDList.place(relx=0.13, rely=0.15, relheight=0.03, relwidth=0.35)

        self.DownloadButton = tk.Button(self, text='Download', anchor='w', justify='left',
                                        command=self.azure_sync)
        self.DownloadButton.place(relx=0.13, rely=0.2, height=30, width=100)

        self.FolderButton = tk.Button(self, text='Open Folder', anchor='w', justify='left',
                                      command=self.open_folder)
        self.FolderButton.place(relx=0.25, rely=0.2, height=30, width=100)

        self.LogText = tk.Text(self, font=("Helvetica", 8))
        self.LogText.place(relx=0.01, rely=0.25, height=550, width=850)
        self.LogText.insert(tk.END, "Logging:\n")
        self.LogText.focus_set()
        self.LogText.focus_lastfor()
        self.LogText.see(tk.END)

        self.progressbar = ttk.Progressbar(self, orient="horizontal", length=900, mode="determinate")
        self.progressbar.pack(side=tk.BOTTOM)

        # self.MethodLabel = tk.Label(self, text='Method:', anchor='w', justify='left')
        # self.MethodLabel.place(relx=0.01, rely=0.35, height=10, width=100)

    def azure_get_directory(self):
        from azure.storage.filedatalake import DataLakeServiceClient
        global PQA_DATA_URL
        global PQA_DATA_CREDENTIAL
        global TEF_DATA_URL
        global TEF_DATA_CREDENTIAL
        service = DataLakeServiceClient(account_url=PQA_DATA_URL,
                                        credential=PQA_DATA_CREDENTIAL)
        file_systems = service.list_file_systems()
        for file_system in file_systems:
            data_list = []
            print("Download File System:", file_system.name)
            file_system_client = service.get_file_system_client(file_system=file_system.name)
            # if "telemetr1y" in file_system.name:
            paths = file_system_client.get_paths()
            for path in paths:
                if not path.is_directory:
                    print("Found Files:", path.name)
                    entry_data = ''
                    entry_decode = ''
                    # filename = path.name.replace("/", "_")
                    # directory_client = file_system_client.get_directory_client(directory=path.name)
                    # file_client = file_system_client.get_file_client(file_path=path.name)

    def get_timestamp(self, datablock):
        match_list = []
        regex_list = ["\"SAMPLING_TM\": \"?(\d*)\"?",
                      "\"REPORT_SBS_TM\": \"?(\d*)\"?",
                      "\"IP_TABLES_TM\": \"?(\d*)\"?",
                      "\"TIMESTAMP_TM\": \"?(\d*)\"?",
                      "\"Timestamp_Sampling\": \"?(\d*)\"?"]
        for re_item in regex_list:
            regex = re.compile(re_item)
            match = regex.findall(datablock)
            if bool(match) == True:
                match_list.append(match)
        return match_list

    def open_folder(self):
        curr_directory = os.getcwd()  # will get current working directory
        name = filedialog.askopenfilename(initialdir=curr_directory, title="Select file",
                                          filetypes=(("excel files", "*.xlsx"), ("all files", "*.*")))
        print(name)

    def download_process(self, download_file, file_type):
        print("Download Files:", download_file)

        entry_data = ''
        entry_decode = ''
        filename = download_file.replace("/", "_")
        # directory_client = file_system_client.get_directory_client(directory=path.name)
        file_client = self.file_system_client.get_file_client(file_path=download_file)
        # file_client = directory_client.get_file_client(file=path.name)
        download = file_client.download_file(max_concurrency=12, timeout=300)
        downloaded_bytes = download.readall().decode("utf-8", errors='ignore')
        # data_list = data_list + downloaded_bytes
        for entry in downloaded_bytes.split("\n"):
            try:
                self.data_list_raw.append(entry)
                if file_type == "telemetry":
                    entry_data = json.loads(entry)['Body']
                    if 'sequenceNumber' not in entry_data:
                        if "UserID" not in entry_data:
                            entry_decode = base64.b64decode(entry_data)
                            entry_data = json.loads(entry_decode.decode("utf-8"))
                        new_dict_data = {}
                        new_dict_data['Data_Type'] = "telemetry"
                        new_dict_data.update(entry_data)
                        self.data_list.append(json.dumps(new_dict_data))
                elif file_type == "twin":
                    if 'properties' in json.loads(entry)['Body']:
                        if "reported" in json.loads(entry)['Body']['properties']:
                            entry_data = json.loads(entry)['Body']['properties']['reported']
                            new_dict_data = {}
                            new_dict_data['Data_Type'] = "twin"
                            new_dict_data.update(entry_data)
                            self.data_list.append(json.dumps(new_dict_data))
                # elif file_type == "log":
                #     if "reported" in json.loads(entry)['Body']['properties']:
                #         entry_data = json.loads(entry)['Body']['properties']['reported']
                #         new_dict_data = {}
                #         new_dict_data['Data_Type'] = "log"
                #         new_dict_data.update(entry_data)
                #         self.data_list.append(json.dumps(new_dict_data))
            except Exception as e:
                print("Error Parser: ", e)
                print("Append RAW Content:", entry)
                self.LogText.insert(tk.END, "Error Parser: Append RAW Content: " + entry + "\r\n")
                self.LogText.see(tk.END)
                self.data_list.append(json.dumps(entry_data))
                continue

    def azure_sync(self):
        from azure.storage.filedatalake import DataLakeServiceClient

        # from multiprocessing.dummy import Pool as ThreadPool
        # from multiprocessing import Manager as Manager
        # from multiprocessing import freeze_support

        # freeze_support()

        file_list = []
        start_poll_date = ''
        end_poll_date = ''
        download = True
        data_list = []
        self.deviceID_list = {}

        source = self.ServerList.get()
        sync_date = self.DateList.get()
        deviceID_query = self.DeviceIDList.get()

        self.progressbar["value"] = 0

        if sync_date == "today":
            sync_date = (time.strftime("%Y/%m/%d", time.localtime()))

        if source == "PQA":
            service = DataLakeServiceClient(account_url=PQA_DATA_URL,
                                            credential=PQA_DATA_CREDENTIAL)
        else:
            service = DataLakeServiceClient(account_url=TEF_DATA_URL,
                                            credential=TEF_DATA_CREDENTIAL)

        file_systems = service.list_file_systems()
        # self.data_list = Manager().list()
        # self.data_list_raw = Manager().list()

        self.data_list = []
        self.data_list_raw = []
        self.LogText.insert(tk.END, "Start Download File:\r\n")
        for file_system in file_systems:
            sync_system = False

            if source == "TEF" and "telemetry" in file_system.name:
                sync_system = True
            elif source == "PQA":
                sync_system = True
            else:
                sync_system = False

            if sync_system == True:
                if "tefdata2-twin" in file_system.name:
                    download_type = "twin"
                elif "tefdata2" in file_system.name:
                    download_type = "telemetry"
                # elif "insights" in file_system.name:
                #     download_type = "log"

                print("Download File System:", file_system.name)
                self.LogText.insert(tk.END, "Download File System: " + file_system.name + "\r\n")
                self.LogText.see(tk.END)

                self.file_system_client = service.get_file_system_client(file_system=file_system.name)

                twin_download_file_list = []
                telemetry_download_file_list = []
                log_download_file_list = []

                paths = self.file_system_client.get_paths()
                for path in paths:
                    download = False
                    if not path.is_directory:
                        if sync_date == "all":
                            download = True
                        else:
                            if sync_date in path.name:
                                download = True
                            else:
                                download = False

                        if download == True:
                            if download_type == "twin":
                                twin_download_file_list.append(path.name)
                            elif download_type == "telemetry":
                                telemetry_download_file_list.append(path.name)
                            # elif download_type == "log":
                            #     log_download_file_list.append(path.name)

                download_telemetry_pool = ""
                thread_count = 0
                # download_telemetry_pool = ThreadPool(5)
                for file in telemetry_download_file_list:
                    self.download_process(file, "telemetry")
                    # download_telemetry_pool=Thread(target=self.download_process,args=(file, "telemetry"))
                    # download_telemetry_pool[thread_count].daemon=True
                    # download_telemetry_pool.start()
                    # download_telemetry_pool.join()
                    # print(dir(download_telemetry_pool))
                    if len(sys.argv) == 1:
                        self.LogText.insert(tk.END, "Download Files: " + file + "\r\n")
                        self.LogText.see(tk.END)
                        self.update()
                        self.update_idletasks()
                        app.update()
                    thread_count += 1

                # for count in range(thread_count):
                # while download_telemetry_pool.is_alive():
                if len(sys.argv) == 1:
                    self.update()
                    self.update_idletasks()
                    app.update()
                self.after(500)

                # download_telemetry_pool.close()

                download_twin_pool = ""
                thread_count = 0
                # download_telemetry_pool = ThreadPool(5)
                for file in twin_download_file_list:
                    self.download_process(file, "twin")
                    # download_twin_pool=Thread(target=self.download_process, args=(file, "twin"))
                    # download_twin_pool[thread_count].daemon = True
                    # download_twin_pool.start()
                    # download_twin_pool.join()
                    # print(dir(download_telemetry_pool))
                    if len(sys.argv) == 1:
                        self.LogText.insert(tk.END, "Download Files: " + file + "\r\n")
                        self.LogText.see(tk.END)
                        self.update()
                        self.update_idletasks()
                        app.update()
                    thread_count += 1

                # for count in range(thread_count):
                # while download_twin_pool.is_alive():
                if len(sys.argv) == 1:
                    self.update()
                    self.update_idletasks()
                    app.update()
                self.after(500)

                thread_count = 0
                # download_telemetry_pool = ThreadPool(5)
                # for file in log_download_file_list:
                #     self.download_process(file, "log")
                #     # download_twin_pool=Thread(target=self.download_process, args=(file, "twin"))
                #     # download_twin_pool[thread_count].daemon = True
                #     # download_twin_pool.start()
                #     # download_twin_pool.join()
                #     # print(dir(download_telemetry_pool))
                #     if len(sys.argv) == 1:
                #         self.LogText.insert(tk.END, "Download Files: " + file + "\r\n")
                #         self.LogText.see(tk.END)
                #         self.update()
                #         self.update_idletasks()
                #         app.update()
                #     thread_count += 1
                # download_twin_pool = ThreadPool(5)
                # for file in twin_download_file_list:
                #    download_twin_pool.apply(self.download_process,args=(file,"twin"))
                #    download_twin_pool.close()
                #    download_twin_pool.join ()

                if len(sys.argv) == 1:
                    self.update()
                    self.update_idletasks()
                    app.update()

        if len(sys.argv) == 1:
            self.progressbar["value"] = 30
            self.progressbar.update()
            self.update()
            self.update_idletasks()
            app.update()

        # Output RAW data
        local_file = open('opah_raw_data.txt', 'w')
        for data_row in self.data_list_raw:
            try:
                local_file.write(data_row + "\n")
            except Exception as e:
                print("Error Write Data:", data_row)
                self.LogText.insert(tk.END, "Error Write Data:" + json.dumps(data_row) + "\r\n")
                self.LogText.see(tk.END)
                continue
        local_file.close()
        print("Output RAW Data Finished")
        self.LogText.insert(tk.END, "Output RAW Data Finished" + "\r\n")
        self.LogText.see(tk.END)

        if len(sys.argv) == 1:
            self.progressbar["value"] = 50
            self.progressbar.update()
            self.update()
            self.update_idletasks()
            app.update()

        # Parser message list
        for entry in self.data_list:
            try:
                deviceID = json.loads(entry)['DeviceId']
                if deviceID not in self.deviceID_list:
                    self.deviceID_list[deviceID] = list()
                self.deviceID_list[deviceID].append([entry])
            except Exception as e:
                print("Parser Exception:", repr(e))
                print("Error Line:", json.dumps(entry))
                self.LogText.insert(tk.END,
                                    "Parser Data_List Exception: " + json.dumps(entry).replace("\\", "") + "\r\n")
                self.LogText.see(tk.END)
                continue
        print("Parser RAW Data Finished")
        self.LogText.insert(tk.END, "Parser RAW Data Finished" + "\r\n")
        self.LogText.see(tk.END)

        if deviceID_query != "all":
            if deviceID_query in self.deviceID_list:
                tmp_data = deepcopy(self.deviceID_list)
                tmp_list = tmp_data[deviceID_query]
                self.deviceID_list = {}
                self.deviceID_list[deviceID_query] = tmp_list
            else:
                messagebox.showinfo("Info",
                                    "Can't Find this " + deviceID_query + " in the list, download all deviceID now, Please wait")
                self.deviceID_list = self.deviceID_list
        else:
            self.deviceID_list = self.deviceID_list

        self.LogText.insert(tk.END, "Generate Excel Report now, please wait......." + "\r\n")
        self.LogText.see(tk.END)

        report_thread = []
        report_thread_id = 0
        task_list = []
        # with ProcessPoolExecutor(max_workers=4) as p:
        #     for deviceID in self.deviceID_list:
        #         task_list.append(p.submit(self.generate_deviceID_file, deviceID))
        #     for task in as_completed(task_list):
        #         if task.done():
        #             print(task.result())

        for deviceID in self.deviceID_list:
            self.generate_deviceID_file(deviceID)
            # report_thread.append(Thread(target=self.generate_deviceID_file,args=(deviceID,)))
            # report_thread[report_thread_id].daemon=True
            # report_thread[report_thread_id].start()
            if len(sys.argv) == 1:
                self.LogText.insert(tk.END, "Generate File ID: " + deviceID + "\r\n")
                self.LogText.see(tk.END)
                self.update()
                self.update_idletasks()
                app.update()
            else:
                print("Generate File ID: ", deviceID)
            report_thread_id += 1

        # for count in range(report_thread_id):
        #    while report_thread[count].is_alive():
        #        if len(sys.argv) == 1:
        #            self.update()
        #            self.update_idletasks()
        #            app.update()
        #        self.after(500)

        if len(sys.argv) == 1:
            self.progressbar["value"] = 100
            self.progressbar.update()
            self.update()
            self.update_idletasks()
            app.update()

        print("Save Data to EXCEL file Finished")
        self.LogText.insert(tk.END, "Save Data to EXCEL file Finished\r\n")
        self.LogText.see(tk.END)

        if len(sys.argv) == 1:
            messagebox.showinfo("Info", "Download Data File Finished")
        else:
            print("Download Data File Finished")

    def generate_deviceID_file(self, deviceID):
        from openpyxl import Workbook
        from openpyxl.styles.alignment import Alignment
        from openpyxl.utils import get_column_letter
        import jsbeautifier
        from datetime import datetime
        sheet_name = {}
        sheet_row = {}

        filename = deviceID + ".xlsx"

        if len(sys.argv) == 1:
            self.progressbar["value"] += 7
            self.update()
            self.update_idletasks()

        try:
            if os.path.exists(filename):
                os.remove(filename)
        except Exception as e:
            self.LogText.insert(tk.END, "Error Remove excel file: " + repr(e) + "\r\n")
            self.LogText.see(tk.END)

        wb = Workbook()
        wb.remove(wb['Sheet'])
        for name in datablock_id:
            column = 1
            if "Watchdog" in name:
                name = name.replace('Watchdog', "WD")
            sheet_name[name] = wb.create_sheet(name)
            sheet_name[name].cell(1, 1).value = "Date/Time"
            column = 2
            for id_name in valid_column:
                sheet_name[name].cell(1, column).value = id_name
                column += 1
            sheet_row[name] = 2

        for entry in self.deviceID_list[deviceID]:
            start_time = datetime.now()
            # print('Handle JSON Entry Time elapsed (hh:mm:ss.ms) {}'.format(start_time))
            # print(entry)

            column = 1
            time_elapsed = datetime.now() - start_time
            # print('1 Time elapsed (hh:mm:ss.ms) {}'.format(time_elapsed))

            try:
                json_data = json.loads(entry[0])
                if "Datablock_id" in json_data:
                    if json_data['Datablock_id'] != "null":
                        for block in json_data['Datablock_id']:
                            time_elapsed = datetime.now() - start_time
                            # print('2 Time elapsed (hh:mm:ss.ms) {}'.format(time_elapsed))

                            if len(sys.argv) == 1:
                                self.update()
                                self.update_idletasks()
                                app.update()

                            column = 1
                            if block in datablock_id:
                                if json_data['Timestamp_Report'] > 0:
                                    time_readable = datetime.fromtimestamp(
                                        json_data['Timestamp_Report'] / 1000)
                                else:
                                    time_readable = datetime.fromtimestamp(int(0))
                                time_readable = time_readable.strftime("%Y-%m-%d %H:%M:%S")

                                # print (json_data['Datablock_id'][block])
                                if "Watchdog" in block:
                                    new_name = block.replace('Watchdog', "WD")
                                else:
                                    new_name = block
                                sheet_name[new_name].cell(sheet_row[new_name], column).value = time_readable
                                sheet_name[new_name].column_dimensions[get_column_letter(column)].width = 20
                                sheet_name[new_name].cell(sheet_row[new_name], column).alignment = Alignment(
                                    wrapText=True,
                                    vertical="top")
                                column += 1
                                for name in json_data:
                                    time_elapsed = datetime.now() - start_time
                                    # print('3 Time elapsed (hh:mm:ss.ms) {}'.format(time_elapsed))

                                    if name in valid_column:
                                        if name == "Datablock_id":
                                            column += 1
                                            # output = json.dumps(json_data[name][block], check_circular=True, allow_nan=True, ensure_ascii=True, indent=1, sort_keys=False)
                                            # output = json.dumps(json_data[name][block], indent=2, sort_keys=False )
                                            time_elapsed = datetime.now() - start_time
                                            # print('4 Time elapsed (hh:mm:ss.ms) {}'.format(time_elapsed))

                                            opts = jsbeautifier.default_options()
                                            opts.indent_size = 1
                                            output = jsbeautifier.beautify(json.dumps(json_data[name][block]), opts)
                                            split_output = output.splitlines()
                                            count = 0

                                            time_elapsed = datetime.now() - start_time
                                            # print('5 Time elapsed (hh:mm:ss.ms) {}'.format(time_elapsed))

                                            new_output = ""
                                            new_column = column
                                            if len(split_output) > 230:
                                                end_count = len(split_output)
                                                # print(split_output)
                                                for line_id in range(len(split_output)):
                                                    new_output = new_output + split_output[line_id] + "\n"
                                                    count = count + 1
                                                    end_count = end_count - 1
                                                    if count >= 230 or end_count == 0:
                                                        sheet_name[new_name].cell(sheet_row[new_name],
                                                                                  new_column).value = str(new_output)
                                                        sheet_name[new_name].cell(sheet_row[new_name],
                                                                                  new_column).alignment = Alignment(
                                                            wrapText=True, vertical="top")
                                                        sheet_name[new_name].column_dimensions[
                                                            get_column_letter(new_column)].width = 40
                                                        # sheet_name[new_name].column_dimensions[str(chr(64 + new_column))].bestFit = True
                                                        # sheet_name[new_name].column_dimensions[str(chr(64 + new_column))].auto_size = True
                                                        new_column = new_column + 1
                                                        count = 0
                                                        new_output = ""
                                            else:
                                                sheet_name[new_name].cell(sheet_row[new_name], column).value = str(
                                                    output)
                                                sheet_name[new_name].column_dimensions[
                                                    get_column_letter(column)].width = 40
                                                sheet_name[new_name].cell(sheet_row[new_name],
                                                                          column).alignment = Alignment(wrapText=True,
                                                                                                        vertical="top")
                                                # sheet_name[new_name].column_dimensions[str(chr(64 + column))].bestFit = True
                                                # sheet_name[new_name].column_dimensions[str(chr(64 + column))].auto_size = True
                                            # if "WD" in new_name:
                                            #    print (output)
                                            match_data = self.get_timestamp(output)
                                            time_list = ""
                                            for match_entry in match_data:
                                                for time_entry in match_entry:
                                                    if time_entry != "":
                                                        if block == "Multicast":
                                                            if json_data['Timestamp_Report'] > 0:
                                                                time_readable = datetime.fromtimestamp(
                                                                    int(time_entry) / 1000)
                                                            else:
                                                                time_readable = datetime.fromtimestamp(int(0))
                                                        else:
                                                            if json_data['Timestamp_Report'] > 0:
                                                                time_readable = datetime.fromtimestamp(
                                                                    int(time_entry) / 1000)
                                                            else:
                                                                time_readable = datetime.fromtimestamp(int(0))
                                                        time_readable = time_readable.strftime("%Y-%m-%d %H:%M:%S")

                                                        time_list = time_list + time_entry + " -> " + time_readable + "\r\n"
                                            sheet_name[new_name].cell(sheet_row[new_name], column - 1).value = str(
                                                time_list)
                                            sheet_name[new_name].column_dimensions[
                                                get_column_letter(column - 1)].width = 40
                                            sheet_name[new_name].cell(sheet_row[new_name],
                                                                      column - 1).alignment = Alignment(
                                                wrapText=True, vertical="top")

                                        else:
                                            if name != "Datablock_Timestamp":
                                                sheet_name[new_name].cell(sheet_row[new_name], column).value = str(
                                                    json_data[name])
                                                sheet_name[new_name].cell(sheet_row[new_name],
                                                                          column).alignment = Alignment(
                                                    wrapText=True, vertical="top")
                                                sheet_name[new_name].column_dimensions[
                                                    get_column_letter(column)].width = 20
                                        column += 1
                                sheet_row[new_name] += 1
                else:
                    print("Error Parser JSON without Datablock_id:", entry)
            except Exception as e:
                print("Output Excel exception:", repr(e))
                print("Error Excel Parser:", entry)
                self.LogText.insert(tk.END, "Error Excel Parser: " + json.dumps(entry) + "\r\n")
                self.LogText.see(tk.END)

                continue
        wb.save(filename)


class twin_update_app(ttk.Frame):
    def __init__(self, top=None):
        super().__init__()

        self.server = ""
        self.device_id = ""
        self.datablock = ""
        self.twin_info = ""
        self.twin_data_detail = ""

        self.InputFrame = tk.Frame(self)
        self.InputFrame.place(relx=0.01, rely=0.02, relheight=0.3, relwidth=0.98)
        self.InputFrame.configure(relief='groove')
        self.InputFrame.configure(borderwidth="2")

        # self.ControlFrame = tk.Frame(self)
        # self.ControlFrame.place(relx=0.01, rely=0.25, relheight=0.1, relwidth=0.98)
        # self.ControlFrame.configure(relief='groove')
        # self.ControlFrame.configure(borderwidth="2")

        self.OutputFrame = tk.Frame(self)
        self.OutputFrame.place(relx=0.01, rely=0.2, relheight=0.85, relwidth=0.98)
        self.OutputFrame.configure(relief='groove')
        self.OutputFrame.configure(borderwidth="2")

        # InputFrame
        self.ServerLabel = tk.Label(self.InputFrame, text='Server:', anchor='w', justify='left')
        self.ServerLabel.place(relx=0.01, rely=0.05, height=15, width=100)

        self.ServerList = ttk.Combobox(self.InputFrame, width=100, values=source)
        self.ServerList.current(0)
        self.ServerList.place(relx=0.13, rely=0.05, relheight=0.1, relwidth=0.35)

        self.DeviceIDLabel = tk.Label(self.InputFrame, text='DeviceID:', anchor='w', justify='left')
        self.DeviceIDLabel.place(relx=0.01, rely=0.2, height=15, width=100)

        self.DeviceIDText = tk.Entry(self.InputFrame, width=100)
        self.DeviceIDText.place(relx=0.13, rely=0.2, relheight=0.1, relwidth=0.35)
        self.DeviceIDText.insert(tk.END, "cc5d4e53225005b5c907bf610e1a7d8c")

        self.DataBlockLabel = tk.Label(self.InputFrame, text='DataBlock:', anchor='w', justify='left')
        self.DataBlockLabel.place(relx=0.01, rely=0.35, height=15, width=100)

        self.DataBlockList = ttk.Combobox(self.InputFrame, width=100, values=datablock_config_id)
        self.DataBlockList.current(0)
        self.DataBlockList.place(relx=0.13, rely=0.35, relheight=0.1, relwidth=0.35)

        self.GetTwinButton = tk.Button(self.InputFrame, text='Get Twin', anchor='w', justify='left',
                                       command=self.get_twin)
        self.GetTwinButton.place(relx=0.55, rely=0.1, height=30, width=100)

        # self.GetDataBlockButton = tk.Button(self.InputFrame, text='Get DataBlock', anchor='w', justify='left',
        #                                    command=self.get_datablock)
        # self.GetDataBlockButton.place(relx=0.13, rely=0.65, height=30, width=100)

        self.UpdateTwinButton = tk.Button(self.InputFrame, text='Update Twin', anchor='w', justify='left',
                                          command=self.update_twin_content)
        self.UpdateTwinButton.place(relx=0.55, rely=0.3, height=30, width=100)

        self.ClearCheck = ttk.Checkbutton(self.InputFrame, text='Clear Desired')
        self.ClearCheck.place(relx=0.7, rely=0.30, height=30, width=200)
        self.ClearCheck.state(['!alternate'])

        # OutputFrame
        self.TwinSourceLabel = tk.Label(self.OutputFrame, text='Reported:', anchor='w', justify='left')
        self.TwinSourceLabel.place(relx=0.01, rely=0.01, height=15, width=100)

        self.TwinSourceText = tk.Text(self.OutputFrame, font=("Helvetica", 9))
        self.TwinSourceText.place(relx=0.01, rely=0.04, height=550, width=400)
        # self.SendText.insert(tk.END, json.dumps({}, indent=4, sort_keys=False))

        self.TwinUpdateLabel = tk.Label(self.OutputFrame, text='Desired:', anchor='w', justify='left')
        self.TwinUpdateLabel.place(relx=0.49, rely=0.01, height=15, width=100)

        self.TwinUpdateText = tk.Text(self.OutputFrame, font=("Helvetica", 9))
        self.TwinUpdateText.place(relx=0.49, rely=0.04, height=550, width=430)

        if len(sys.argv) == 1:
            self.read_config()

    def get_twin(self):
        self.server = self.ServerList.get()
        self.device_id = self.DeviceIDText.get()
        self.datablock = self.DataBlockList.get()
        self.TwinSourceText.delete("1.0", tk.END)
        self.TwinUpdateText.delete("1.0", tk.END)

        if self.server == "PQA":
            connection = PQA_CONNECTION_STRING
        else:
            connection = TEF_CONNECTION_STRING

        try:
            self.registry_manager = IoTHubRegistryManager(connection)
            self.twin_info = self.registry_manager.get_twin(self.device_id)
            if self.datablock in self.twin_info.properties.reported["Datablock_id"]:
                self.twin_data_raw = self.twin_info.properties.reported["Datablock_id"][self.datablock]
                self.twin_data_text = "{ \"Datablock_id\": { " + "\"" + self.datablock + "\": " + \
                                      json.dumps(
                                          self.twin_info.properties.reported["Datablock_id"][self.datablock]) + " } }"

                self.twin_data_json = json.loads(self.twin_data_text)
                self.TwinSourceText.insert(tk.END, json.dumps(self.twin_data_json, indent=4, sort_keys=False))

                if len(sys.argv) == 1:
                    self.output_twin()
                else:
                    print("Output:")
                    print(json.dumps(self.twin_data_json, indent=4, sort_keys=False))

            else:
                self.TwinSourceText.insert(tk.END, "Error: No Datablock found in Reported Twin")
                self.TwinUpdateText.insert(tk.END, "Error: No Datablock found in Reported Twin")
                # print (self.twin_data_detail)
            if len(sys.argv) == 1:
                messagebox.showinfo("Info", "Get Twin Finished")
            else:
                print("Get Twin Finished")
        except Exception as e:
            if len(sys.argv) == 1:
                messagebox.showerror("Error Get Twin:", repr(e))
            else:
                print("Error Get Twin:", repr(e))

    def read_config(self):
        self.config_table = []
        try:
            csv_file = open(".\\setting\\DataBlocks_Config__Remove_Info.csv", "r")
            dict_reader = csv.DictReader(csv_file)
            for row in dict_reader:
                self.config_table.append(dict(row))
        except Exception as e:
            print("Error: ", repr(e))

        # print(self.config_table)

    def remove_keys(self, obj, rubbish):
        if isinstance(obj, dict):
            obj = {
                key: self.remove_keys(value, rubbish)
                for key, value in obj.items()
                if key not in rubbish}
        elif isinstance(obj, list):
            obj = [self.remove_keys(item, rubbish)
                   for item in obj
                   if item not in rubbish]
        return obj

    def clean_dict(self, obj_json, bad_key):
        """
        This method scrolls the entire 'obj' to delete every key for which the 'callable' returns
        True

        :param obj: a dictionary or a list of dictionaries to clean
        :param func: a callable that takes a key in argument and return True for each key to delete
        """
        if isinstance(obj_json, dict):
            # the call to `list` is useless for py2 but makes
            # the code py2/py3 compatible
            for key in list(obj_json.keys()):
                if key == bad_key:
                    del obj_json[key]
                else:
                    self.clean_dict(obj_json[key], bad_key)
        elif isinstance(obj_json, list):
            for i in reversed(range(len(obj_json))):
                if obj_json[i] == bad_key:
                    del obj_json[i]
                else:
                    self.clean_dict(obj_json[i], bad_key)
        else:
            # neither a dict nor a list, do nothing
            pass
        return obj_json

    def output_twin(self):

        self.twin_data_output_json = deepcopy(self.twin_data_json)
        for row in self.config_table:
            if row['DataBlock'] == self.datablock:
                self.twin_data_output_json = self.clean_dict(self.twin_data_output_json, row['Name'])

        output_text = json.dumps(self.twin_data_output_json, indent=4, sort_keys=False)

        if 'selected' in self.ClearCheck.state():
            regex = re.compile(r"\": [+-]?[0-9]+\.[0-9]+", re.IGNORECASE)
            output_text = regex.sub("\": null", output_text)
            regex = re.compile(r"\": [0-9]+", re.IGNORECASE)
            output_text = regex.sub("\": null", output_text)
            regex = re.compile(r"\": \".*\"", re.IGNORECASE)
            output_text = regex.sub("\": null", output_text)
            regex = re.compile(r"\": true", re.IGNORECASE)
            output_text = regex.sub("\": null", output_text)
            regex = re.compile(r"\": false", re.IGNORECASE)
            output_text = regex.sub("\": null", output_text)

        self.TwinUpdateText.insert(tk.END, output_text)

    def get_datablock(self):
        print("test")

    def update_twin_content(self):
        # Update twin
        from azure.iot.hub.models import Twin, TwinProperties

        if self.server == "PQA":
            connection = PQA_CONNECTION_STRING
        else:
            connection = TEF_CONNECTION_STRING

        twin_patch = Twin()
        self.registry_manager = IoTHubRegistryManager(connection)
        try:
            output_twin = json.loads(self.TwinUpdateText.get("1.0", tk.END))
            twin_patch.properties = TwinProperties(desired=output_twin)
            updated_module_twin = self.registry_manager.update_twin(self.device_id, twin_patch, None)
            # print(updated_module_twin)
            if len(sys.argv) == 1:
                messagebox.showinfo("Info", "Update Twin Finished")
            else:
                print("Update Twin Finished")
        except Exception as e:
            if len(sys.argv) == 1:
                messagebox.showerror("Error Update Twin", repr(e))
            else:
                print("Error Update Twin", repr(e))

        # for i in range(2):
        #    self.update()
        #    self.after(1000)
        #    self.update_idletasks()
        #    self.update()
        # self.get_twin()


if __name__ == '__main__':
    freeze_support()
    start_time = time.time()
    if len(sys.argv) > 1:
        parser = ArgumentParser()
        parser.add_argument("-a", "--action", dest="action")
        parser.add_argument("-s", "--server", dest="server")
        parser.add_argument("-d", "--date", dest="date")
        parser.add_argument("-D", "--deviceID", dest="deviceID")
        parser.add_argument("-m", "--method", dest="method")
        parser.add_argument("-p", "--paload", dest="payload")
        parser.add_argument("-b", "--datablock", dest="datablock")
        parser.add_argument("-r", "--request", dest="request")
        args = parser.parse_args()
        # print("--- %s seconds ---" % (time.time() - start_time))
        if (args.action == "download_telemetry"):
            if (args.server == "" or args.date == "" or args.deviceID == ""):
                print("Please input -s -d -D parameter!!")
            else:
                # print("--- %s seconds ---" % (time.time() - start_time))
                download_process = datalake_app()
                download_process.pack_forget()
                download_process.ServerList.delete(0, tk.END)
                download_process.ServerList.insert(tk.END, args.server)
                download_process.DateList.delete(0, tk.END)
                download_process.DateList.insert(tk.END, args.date)
                download_process.DeviceIDList.delete(0, tk.END)
                download_process.DeviceIDList.insert(tk.END, args.deviceID)
                # print("--- %s seconds ---" % (time.time() - start_time))
                download_process.azure_sync()
                # print("--- %s seconds ---" % (time.time() - start_time))
        elif (args.action == "direct_method"):
            if (args.server == "" or args.method == "" or args.payload == "" or args.deviceID == ""):
                print("Please input -s -m -p -D parameter!!")
            else:
                # print("--- %s seconds ---" % (time.time() - start_time))
                payload_file = open(args.payload, "r")
                payload_content = json.loads(payload_file.read())
                payload_content = json.dumps(payload_content)
                payload_file.close()
                # print("--- %s seconds ---" % (time.time() - start_time))
                send_direct_method("", args.server, args.deviceID, args.method, payload_content)
                # print("--- %s seconds ---" % (time.time() - start_time))
        elif (args.action == "twin_update"):
            if (
                    args.server == "" or args.datablock == "" or args.request == "" or args.payload == "" or args.deviceID == ""):
                print("Please input -s -b -r -p -D parameter!!")
            else:
                if (args.request == "get"):
                    update_process = twin_update_app()
                    update_process.pack_forget()
                    update_process.ServerList.delete(0, tk.END)
                    update_process.ServerList.insert(tk.END, args.server)
                    update_process.DataBlockList.delete(0, tk.END)
                    update_process.DataBlockList.insert(tk.END, args.datablock)
                    update_process.DeviceIDText.delete(0, tk.END)
                    update_process.DeviceIDText.insert(tk.END, args.deviceID)
                    update_process.get_twin()
                elif (args.request == "update"):
                    update_process = twin_update_app()
                    update_process.pack_forget()
                    update_process.ServerList.delete(0, tk.END)
                    update_process.ServerList.insert(tk.END, args.server)
                    update_process.server = args.server
                    update_process.DataBlockList.delete(0, tk.END)
                    update_process.DataBlockList.insert(tk.END, args.datablock)
                    update_process.DeviceIDText.delete(0, tk.END)
                    update_process.DeviceIDText.insert(tk.END, args.deviceID)
                    update_process.device_id = args.deviceID

                    payload_file = open(args.payload, "r")
                    payload_content = json.loads(payload_file.read())
                    payload_content = json.dumps(payload_content)
                    payload_file.close()

                    update_process.TwinUpdateText.delete("1.0", tk.END)
                    update_process.TwinUpdateText.insert(tk.END, payload_content)
                    update_process.update_twin_content()

    else:
        app = main_app()
        app.mainloop()
        # sys.exit(0)
